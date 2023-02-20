from django.shortcuts import render
from openpyxl import load_workbook
from .forms import UploadFileForm


# Create your views here.


def index(request):
    return render(request, 'report_generator/index.html')


"""
    row 0: Machine Number
    row 1: Operation Hours
    row 2: Claim Created at
    row 3: BGZ Master Structure
    row 4: BGZ Material Text
    row 5: BGZ Material Text
    row 6 - 11: Failure Long Text ok
    row 12 - 17: Long Text Diagnostic ok
    row 18 - 23: Long Text Remedy
    row 24 - 28: Long Text Reason
    row 29 - 36: Long Text Comment
"""


def variable_view(request):
    data = []
    if request.method == 'POST':
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            file = request.FILES['file']
            wb = load_workbook(filename=file)
            ws = wb.active
            for row in ws.iter_rows(values_only=True):
                if row[0] is not None:
                    failure_long_text = str(row[6]) + str(row[7]) + str(row[8]) \
                                        + str(row[9]) + str(row[10]) + str(row[11])
                    diagnostic_long_text = str(row[12]) + str(row[13]) + str(row[14]) \
                                                        + str(row[15]) + str(row[16]) + str(row[17])
                    remedy_long_text = str(row[18]) + str(row[19]) + str(row[20]) \
                                                    + str(row[21]) + str(row[22]) + str(row[23])
                    reason_long_text = str(row[24]) + str(row[25]) + str(row[26]) \
                                                    + str(row[27]) + str(row[28])
                    comment_long_text = str(row[29]) + str(row[30]) + str(row[31]) \
                                                     + str(row[32]) + str(row[33]) + str(row[31]) \
                                                     + str(row[32]) + str(row[33])
                    print(failure_long_text, diagnostic_long_text, remedy_long_text, reason_long_text, comment_long_text)

    return render(request, 'report_generator/variable.html')
